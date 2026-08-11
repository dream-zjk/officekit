"""Import tabular data (CSV / TSV) into an Excel worksheet."""
from __future__ import annotations

import csv
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook


def import_csv(
    out: str | Path,
    sheet_title: str,
    csv_path: str | Path,
    header: bool = False,
    delimiter: Optional[str] = None,
) -> Path:
    """Write *csv_path* into worksheet *sheet_title* of the xlsx at *out*.

    If *out* already exists the sheet is added/overwritten inside it; otherwise a
    new workbook is created. ``header`` is accepted for compatibility (the first
    row is always written as data). TSV is auto-detected by extension.
    """
    csv_path = Path(csv_path)
    delim = delimiter or ("\t" if csv_path.name.lower().endswith(".tsv") else ",")
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f, delimiter=delim))

    out = Path(out)
    if out.exists():
        wb = load_workbook(out)
        if sheet_title in wb.sheetnames:
            del wb[sheet_title]
        ws = wb.create_sheet(title=sheet_title)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title

    start = 0
    if header and rows:
        ws.append(rows[0])
        start = 1
    for row in rows[start:]:
        ws.append(row)

    wb.save(str(out))
    return out
